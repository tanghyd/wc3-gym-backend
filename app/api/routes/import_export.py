import logging
from io import BytesIO
from typing import Annotated, Any

import openpyxl
from fastapi import APIRouter, Depends, File, Response, UploadFile

from app.api.deps import (
    FantasyBetServiceDep,
    FantasyTeamServiceDep,
    MatchServiceDep,
    SeasonServiceDep,
    SeriesServiceDep,
    TeamServiceDep,
    UserServiceDep,
    require_admin,
)
from app.core.exceptions import BadRequestError
from app.core.query import QueryUtil
from app.models.responses import Message
from app.services.fantasy_import import (
    import_fantasy_bets_workbook,
    import_fantasy_teams_workbook,
)
from app.services.season_import import import_season_workbook

logger = logging.getLogger(__name__)

router = APIRouter(tags=["import export"])

BET_PAGE = 500  # how many bets the export reads per statement


# import export endpoints
@router.post("/import", dependencies=[Depends(require_admin)])
def import_season(
    file: Annotated[UploadFile | None, File()] = None,
    create_new: str = "false",
    score_system: str | None = None,
) -> dict[str, Any]:
    """Import complete season data from Excel.

    Imports ALL season data (season, maps, teams, players, matches, series)
    from Excel file. score_system overrides the one the workbook carries.
    """
    if file is None:
        raise BadRequestError("No file part")

    if file.filename == "" or not file.filename.endswith((".xlsx", ".xls")):
        raise BadRequestError("No selected file or invalid file type")

    imported = import_season_workbook(
        file.file.read(), create_new.lower() == "true", score_system
    )

    return {
        "message": "Season imported successfully",
        "season_id": imported.id,
        "season_name": imported.name,
        "duplicate_bets_skipped": imported.duplicate_bets,
    }


@router.post("/export", dependencies=[Depends(require_admin)])
def export_season(
    season_service: SeasonServiceDep,
    team_service: TeamServiceDep,
    user_service: UserServiceDep,
    match_service: MatchServiceDep,
    series_service: SeriesServiceDep,
    fantasy_team_service: FantasyTeamServiceDep,
    fantasy_bet_service: FantasyBetServiceDep,
    season_id: int,
) -> Response:
    """Export one season as an Excel workbook of ten sheets.

    The workbook holds the season row, its maps, teams, rostered players,
    matches, series, fantasy teams, fantasy team players, fantasy bets and
    the fantasy users who are on no roster.
    """
    # get raises NotFoundError, which answers 404
    season = season_service.get(season_id)

    # write_only keeps one row in memory at a time instead of the whole sheet
    workbook = openpyxl.Workbook(write_only=True)

    # ===== Sheet 1: Season Metadata =====
    season_sheet = workbook.create_sheet(title="Season")
    season_sheet.append(
        [
            "ID",
            "Name",
            "Number of Weeks",
            "Series Per Week",
            "Pick Ban",
            "Start Date",
            "End Date",
            "Discord Role",
            "Score System",
        ]
    )
    season_sheet.append(
        [
            season.id,
            season.name,
            season.number_weeks,
            season.series_per_week,
            season.pick_ban or "",
            season.start_date.strftime("%Y-%m-%d") if season.start_date else "",
            season.end_date.strftime("%Y-%m-%d") if season.end_date else "",
            season.discordRole or "",
            season.score_system,
        ]
    )

    # ===== Sheet 2: Maps =====
    maps_sheet = workbook.create_sheet(title="Maps")
    maps_sheet.append(["ID", "Name", "Shortname", "Image URL"])
    if season.maps:
        for map_obj in season.maps:
            maps_sheet.append(
                [
                    map_obj.id,
                    map_obj.name,
                    map_obj.shortname,
                    map_obj.image or "",
                ]
            )

    # ===== Sheet 3: Teams =====
    teams_sheet = workbook.create_sheet(title="Teams")
    teams_sheet.append(["ID", "Name", "Long Name", "Discord Role", "Icon URL"])
    season_teams = team_service.get_teams_season(season_id)
    for team in season_teams:
        teams_sheet.append(
            [
                team.id,
                team.name,
                team.long_name or "",
                team.discord_role or "",
                "",  # no model carries an icon URL
            ]
        )

    # ===== Sheet 4: Players =====
    players_sheet = workbook.create_sheet(title="Players")
    players_sheet.append(
        [
            "ID",
            "Name",
            "Battle Tag",
            "Discord Tag",
            "Discord ID",
            "Race",
            "MMR",
            "Country",
            "Fantasy Tier",
            "Team ID",
        ]
    )
    roster_user_ids = set()
    for team in season_teams:
        players = team.player_by_season.get(season_id, [])
        for user in players:
            roster_user_ids.add(user.id)
            players_sheet.append(
                [
                    user.id,
                    user.name,
                    user.battleTag,
                    user.discordTag,
                    user.discordId or "",
                    user.race,
                    user.mmr or "",
                    user.country or "",
                    user.fantasy_tier or "",
                    team.id,
                ]
            )

    # ===== Sheet 5: Matches =====
    matches_sheet = workbook.create_sheet(title="Matches")
    matches_sheet.append(
        [
            "ID",
            "Team1 ID",
            "Team2 ID",
            "Season ID",
            "Playday",
            "Team1 Score",
            "Team2 Score",
            "Fixed Map ID",
            "Date Frame",
        ]
    )
    q_string = f"season_id=={season_id}"
    query = QueryUtil.parse_query(q_string)
    if query and query.elementA:
        all_matches = match_service.search(query)
        for match in all_matches:
            matches_sheet.append(
                [
                    match.id,
                    match.team1.id,
                    match.team2.id,
                    match.season.id,
                    match.playday,
                    match.team1_score or "",
                    match.team2_score or "",
                    match.fixed_map.id if match.fixed_map else "",
                    match.date_frame or "",
                ]
            )

    # ===== Sheet 6: Series =====
    series_sheet = workbook.create_sheet(title="Series")
    series_sheet.append(
        [
            "ID",
            "Match ID",
            "Player1 ID",
            "Player2 ID",
            "Player1 Score",
            "Player2 Score",
            "Player1 Points",
            "Player2 Points",
            "Host Player ID",
            "Date Time",
            "Caster",
            "Is Fantasy Match",
        ]
    )
    for match in all_matches if "all_matches" in locals() else []:
        q_string = f"match_id=={match.id}"
        query = QueryUtil.parse_query(q_string)
        if query and query.elementA:
            series_list = series_service.search(query)
            for series in series_list:
                date_time_str = (
                    series.date_time.strftime("%Y-%m-%d %H:%M:%S")
                    if series.date_time
                    else ""
                )
                series_sheet.append(
                    [
                        series.id,
                        series.match.id,
                        series.player1.id,
                        series.player2.id,
                        series.player1_score
                        if series.player1_score is not None
                        else "",
                        series.player2_score
                        if series.player2_score is not None
                        else "",
                        series.player1_points or "",
                        series.player2_points or "",
                        series.host_player_id,
                        date_time_str,
                        series.caster or "",
                        series.is_fantasy_match or False,
                    ]
                )

    # The captains, drafted players and bettors, for the Fantasy Users sheet below
    fantasy_user_ids: set[int | None] = set()

    # ===== Sheet 7: Fantasy Teams =====
    fantasy_teams_sheet = workbook.create_sheet(title="Fantasy Teams")
    fantasy_teams_sheet.append(
        [
            "ID",
            "Name",
            "Season ID",
            "Captain ID",
            "Drafted Team ID",
            "Drafted Race",
            "Player Points",
            "Bench Points",
            "Team Points",
            "Race Points",
            "Bet Points",
            "Total Points",
        ]
    )
    q_string = f"season_id=={season_id}"
    query = QueryUtil.parse_query(q_string)
    if query and query.elementA:
        fantasy_teams, _ = fantasy_team_service.search(query)
        for fteam in fantasy_teams:
            fantasy_user_ids.add(fteam.captain_id)
            fantasy_teams_sheet.append(
                [
                    fteam.id,
                    fteam.name,
                    fteam.season_id,
                    fteam.captain_id,
                    fteam.drafted_team_id or "",
                    fteam.drafted_race or "",
                    fteam.player_points or 0,
                    fteam.bench_points or 0,
                    fteam.team_points or 0,
                    fteam.race_points or 0,
                    fteam.bet_points or 0,
                    fteam.total_points or 0,
                ]
            )

    # ===== Sheet 8: Fantasy Team Players (many-to-many) =====
    fantasy_players_sheet = workbook.create_sheet(title="Fantasy Team Players")
    fantasy_players_sheet.append(["Fantasy Team ID", "Player ID"])
    for fteam in fantasy_teams if "fantasy_teams" in locals() else []:
        if fteam.drafted_players:
            for player in fteam.drafted_players:
                fantasy_user_ids.add(player.id)
                fantasy_players_sheet.append([fteam.id, player.id])

    # ===== Sheet 9: Fantasy Bets =====
    fantasy_bets_sheet = workbook.create_sheet(title="Fantasy Bets")
    fantasy_bets_sheet.append(
        [
            "ID",
            "Season ID",
            "Series ID",
            "User ID",
            "Winner ID",
            "Bet Points",
            "Bet Result",
        ]
    )
    if query and query.elementA:
        # A season holds the most bets of anything here, so it is read by page
        offset = 0
        while True:
            page, _ = fantasy_bet_service.search(query, limit=BET_PAGE, offset=offset)
            for fbet in page:
                fantasy_user_ids.add(fbet.user_id)
                fantasy_user_ids.add(fbet.winner_id)
                fantasy_bets_sheet.append(
                    [
                        fbet.id,
                        fbet.season_id,
                        fbet.series_id,
                        fbet.user_id,
                        fbet.winner_id,
                        fbet.bet_points,
                        fbet.bet_result or "",
                    ]
                )
            if len(page) < BET_PAGE:
                break
            offset += BET_PAGE

    # ===== Sheet 10: Fantasy Users =====
    # The users of the sheets above who are on no roster, so not in Players
    fantasy_users_sheet = workbook.create_sheet(title="Fantasy Users")
    fantasy_users_sheet.append(
        ["ID", "Name", "Battle Tag", "Discord Tag", "Discord ID"]
    )
    for user in user_service.find_by_ids(fantasy_user_ids - roster_user_ids):
        fantasy_users_sheet.append(
            [
                user.id,
                user.name or "",
                user.battleTag,
                user.discordTag or "",
                user.discordId or "",
            ]
        )

    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)

    download_name = f"{season.name.replace(' ', '_')}_export.xlsx"
    return Response(
        content=excel_stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{download_name}"'},
    )


def _workbook_bytes(file: UploadFile | None) -> bytes:
    """The bytes of an uploaded workbook."""
    if file is None:
        raise BadRequestError("No file part")
    if file.filename == "":
        raise BadRequestError("No selected file")
    if not file.filename.endswith((".xlsx", ".xls")):
        raise BadRequestError("File type not allowed")
    return file.file.read()


@router.post("/fantasy/import/teams", dependencies=[Depends(require_admin)])
def import_fantasy_teams(
    file: Annotated[UploadFile | None, File()] = None,
    season_id: str | None = None,
    season_name: str | None = None,
) -> Message:
    """Import a xlsx with the information for a GNL fantasy season.

    Updates the database based on the import sheet.
    """
    import_fantasy_teams_workbook(
        _workbook_bytes(file), int(season_id) if season_id else None, season_name
    )
    return Message(message="File uploaded successfully and data inserted into database")


@router.post("/fantasy/import/bets", dependencies=[Depends(require_admin)])
def import_fantasy_bets(
    file: Annotated[UploadFile | None, File()] = None,
    season_id: str | None = None,
    season_name: str | None = None,
) -> Message:
    """Import a xlsx with the information for a GNL fantasy season.

    Updates the database based on the import sheet.
    """
    import_fantasy_bets_workbook(
        _workbook_bytes(file), int(season_id) if season_id else None, season_name
    )
    return Message(message="File uploaded successfully and data inserted into database")
